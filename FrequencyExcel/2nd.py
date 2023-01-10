import openpyxl as op

word = input("Enter a word : ")

# frequency count 
freuency = 0

# loading the work book
workbook = op.load_workbook("FrequencyExcel\SampleData.xlsx")

# Loading the sheet in which the data is stored
sheet = workbook['sheet1']

# # Traversing the row in Excel sheet
# for row in range(1, sheet.max_row+1):
    
#     # For Each Row Traversing the columns in Excel sheet
#     for col in range(1, sheet.max_column+1):
        
#         # Printing the row and column value with index
#         print("Value at row ",row," and column ",col," is : ", sheet.cell(row,col).value )
        
#         # Deciding current cell has same value as word or not
#         if(sheet.cell(row,col).value == word):
            
#             # Incrementing the count or frequency by 1.
#             freuency += 1
            
#     print("\n")
        
# print("Frequency of the word ",word," ",freuency)

for row in range(1, sheet.max_row+1):
    if(sheet.cell(row,5).value == word):
        freuency += 1
        
print("Frequency of the word in Column 5 ",word," ",freuency)



